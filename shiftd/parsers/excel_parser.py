"""Parse Excel (.xlsx) into TableModel. Requires optional dependency: openpyxl."""

from __future__ import annotations

from pathlib import Path

from shiftd.parsers.registry import register_parser
from shiftd.schema import TableModel


@register_parser("xlsx")
@register_parser("excel")
class ExcelParser:
    """Read the first sheet of an Excel file into TableModel."""

    def parse(self, source: Path | str) -> TableModel:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "Excel support requires optional dependency: uv add 'shiftd[excel]'"
            ) from e
        path = Path(source)
        if not path.exists():
            raise FileNotFoundError(str(path))
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            wb.close()
            return TableModel(columns=[], rows=[])
        columns = [str(c) for c in header]
        rows = [{columns[i]: cell for i, cell in enumerate(row)} for row in rows_iter]
        wb.close()
        if not rows:
            return TableModel(columns=columns, rows=[])
        return TableModel(columns=columns, rows=rows)
